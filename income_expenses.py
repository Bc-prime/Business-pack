import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================
# CONFIGURACIÓN — El cliente solo edita esto
# ============================================
ARCHIVO = "finanzas.xlsx"
NOMBRE_NEGOCIO = "Mi Negocio"
# ============================================

# 1. Leer archivo
df = pd.read_excel(ARCHIVO)
df["Date"] = pd.to_datetime(df["Date"])
df["Month"] = df["Date"].dt.strftime("%B")
df["Month_Num"] = df["Date"].dt.month
print("✅ File loaded successfully")

# 2. Separar ingresos y gastos
ingresos = df[df["Type"] == "Income"]
gastos = df[df["Type"] == "Expense"]

# 3. Totales generales
total_ingresos = ingresos["Amount"].sum()
total_gastos = gastos["Amount"].sum()
balance = total_ingresos - total_gastos
margen = (balance / total_ingresos * 100).round(2) if total_ingresos > 0 else 0

# 4. Por categoría
ingresos_cat = ingresos.groupby("Category")["Amount"].sum().sort_values(ascending=False)
gastos_cat = gastos.groupby("Category")["Amount"].sum().sort_values(ascending=False)

# 5. Por mes
por_mes = df.groupby(["Month_Num", "Month", "Type"])["Amount"].sum().reset_index()
meses_ingresos = por_mes[por_mes["Type"] == "Income"].sort_values("Month_Num")
meses_gastos = por_mes[por_mes["Type"] == "Expense"].sort_values("Month_Num")

# 6. Resumen terminal
print("\n" + "="*55)
print(f"     💰 INCOME VS EXPENSES — {NOMBRE_NEGOCIO}")
print("="*55)
print(f"\n  💚 Total Income:    ${total_ingresos:,.2f}")
print(f"  🔴 Total Expenses:  ${total_gastos:,.2f}")
print(f"  {'💰' if balance >= 0 else '🚨'} Balance:          ${balance:,.2f}")
print(f"  📊 Profit Margin:   {margen}%")

print("\n  📈 Income by Category:")
for cat, amt in ingresos_cat.items():
    print(f"    {cat}: ${amt:,.2f}")

print("\n  📉 Expenses by Category:")
for cat, amt in gastos_cat.items():
    print(f"    {cat}: ${amt:,.2f}")

print("\n  📅 Monthly Summary:")
for _, row in meses_ingresos.iterrows():
    mes_gasto = meses_gastos[meses_gastos["Month"] == row["Month"]]["Amount"].sum()
    mes_balance = row["Amount"] - mes_gasto
    emoji = "✅" if mes_balance >= 0 else "🚨"
    print(f"    {row['Month']}: Income ${row['Amount']:,.2f} | Expenses ${mes_gasto:,.2f} | {emoji} ${mes_balance:,.2f}")
print("="*55)

# 7. Gráficos
fig, axes = plt.subplots(2, 2, figsize=(14, 10))
fig.suptitle(f"Income vs Expenses — {NOMBRE_NEGOCIO}", fontsize=14, fontweight="bold")

# Gráfico 1 — Balance general
labels = ["Income", "Expenses", "Balance"]
values = [total_ingresos, total_gastos, balance]
colors = ["#16A34A", "#DC2626", "#3B82F6"]
bars = axes[0, 0].bar(labels, values, color=colors, edgecolor="black")
axes[0, 0].set_title("General Balance")
axes[0, 0].set_ylabel("Amount ($)")
for bar, val in zip(bars, values):
    axes[0, 0].text(bar.get_x() + bar.get_width()/2, bar.get_height() + 50,
                    f"${val:,.0f}", ha="center", fontsize=9, fontweight="bold")

# Gráfico 2 — Gastos por categoría
gastos_cat.plot(kind="pie", ax=axes[0, 1], autopct="%1.1f%%", startangle=90)
axes[0, 1].set_title("Expenses by Category")
axes[0, 1].set_ylabel("")

# Gráfico 3 — Tendencia mensual
if not meses_ingresos.empty and not meses_gastos.empty:
    axes[1, 0].plot(meses_ingresos["Month"], meses_ingresos["Amount"],
                    marker="o", color="#16A34A", linewidth=2, label="Income")
    axes[1, 0].plot(meses_gastos["Month"], meses_gastos["Amount"],
                    marker="o", color="#DC2626", linewidth=2, label="Expenses")
    axes[1, 0].set_title("Monthly Trend")
    axes[1, 0].set_ylabel("Amount ($)")
    axes[1, 0].tick_params(axis="x", rotation=45)
    axes[1, 0].legend()

# Gráfico 4 — Ingresos por categoría
ingresos_cat.plot(kind="barh", ax=axes[1, 1], color="#16A34A", edgecolor="black")
axes[1, 1].set_title("Income by Category")
axes[1, 1].set_xlabel("Amount ($)")

plt.tight_layout()
plt.savefig("financial_report.png", dpi=150)
plt.show()
print("\n📈 Charts saved as financial_report.png")

# 8. Guardar Excel
wb = Workbook()
ws = wb.active
ws.title = "Financial Report"

DARK = "1E293B"
GREEN = "DCFCE7"
RED = "FEE2E2"
BLUE = "DBEAFE"
WHITE = "FFFFFF"
GRAY = "F1F5F9"

# Título
ws.merge_cells("A1:D1")
t = ws["A1"]
t.value = f"FINANCIAL REPORT — {NOMBRE_NEGOCIO}"
t.font = Font(bold=True, color=WHITE, size=13)
t.fill = PatternFill("solid", fgColor=DARK)
t.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# KPIs
kpis = [
    ("💚 Total Income", f"${total_ingresos:,.2f}", GREEN),
    ("🔴 Total Expenses", f"${total_gastos:,.2f}", RED),
    ("💰 Balance", f"${balance:,.2f}", BLUE),
    ("📊 Profit Margin", f"{margen}%", GRAY),
]
for i, (label, value, bg) in enumerate(kpis):
    row = i + 2
    cell_label = ws.cell(row=row, column=1, value=label)
    cell_label.font = Font(bold=True, size=10)
    cell_label.fill = PatternFill("solid", fgColor=bg)
    cell_label.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

    cell_value = ws.cell(row=row, column=2, value=value)
    cell_value.font = Font(bold=True, size=10)
    cell_value.fill = PatternFill("solid", fgColor=bg)
    cell_value.alignment = Alignment(horizontal="center", vertical="center")

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 18

wb.save("financial_report.xlsx")
print("💾 Report saved as financial_report.xlsx")