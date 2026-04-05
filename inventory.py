import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================
# CONFIGURACIÓN — El cliente solo edita esto
# ============================================
ARCHIVO_INVENTARIO = "inventario.xlsx"
STOCK_MINIMO = 10  # Alerta si hay menos de X unidades
NOMBRE_NEGOCIO = "Mi Negocio"
# ============================================

# 1. Leer inventario
df = pd.read_excel(ARCHIVO_INVENTARIO)
print("✅ Inventario cargado correctamente")

# 2. Detectar productos con stock bajo
stock_bajo = df[df["Stock"] <= STOCK_MINIMO].copy()
stock_ok = df[df["Stock"] > STOCK_MINIMO].copy()

# 3. Ordenar por stock de menor a mayor
stock_bajo = stock_bajo.sort_values("Stock")

# 4. Resumen en terminal
print("\n" + "="*50)
print(f"     📦 INVENTORY ALERT — {NOMBRE_NEGOCIO}")
print("="*50)
print(f"\n  Total products: {len(df)}")
print(f"  ✅ Stock OK: {len(stock_ok)}")
print(f"  🚨 Low stock alerts: {len(stock_bajo)}")

if len(stock_bajo) > 0:
    print(f"\n  🚨 PRODUCTS TO REORDER:")
    for _, row in stock_bajo.iterrows():
        print(f"  ⚠️  {row['Product']} — {row['Stock']} units left")
else:
    print("\n  ✅ All products have sufficient stock!")

print("="*50)

# 5. Generar reporte Excel con colores
wb = Workbook()
ws = wb.active
ws.title = "Inventory Report"

# Colores
RED = "FEE2E2"
GREEN = "DCFCE7"
DARK = "1E293B"
WHITE = "FFFFFF"
ORANGE = "FED7AA"

# Título
ws.merge_cells("A1:E1")
title = ws["A1"]
title.value = f"INVENTORY ALERT — {NOMBRE_NEGOCIO}"
title.font = Font(bold=True, color=WHITE, size=13)
title.fill = PatternFill("solid", fgColor=DARK)
title.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

# Resumen
ws.merge_cells("A2:E2")
summary = ws["A2"]
summary.value = f"Total: {len(df)} products  |  ✅ OK: {len(stock_ok)}  |  🚨 Low stock: {len(stock_bajo)}"
summary.font = Font(size=10, color=DARK)
summary.fill = PatternFill("solid", fgColor="F1F5F9")
summary.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[2].height = 20

# Headers
headers = ["Product", "Category", "Stock", "Min. Stock", "Status"]
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.fill = PatternFill("solid", fgColor="3B82F6")
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[3].height = 22

# Datos
all_products = pd.concat([stock_bajo, stock_ok])
for i, (_, row) in enumerate(all_products.iterrows()):
    r = i + 4
    is_low = row["Stock"] <= STOCK_MINIMO
    bg = RED if is_low else GREEN
    status = "⚠️ REORDER NOW" if is_low else "✅ OK"

    data = [row["Product"], row["Category"], row["Stock"], STOCK_MINIMO, status]
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill = PatternFill("solid", fgColor=bg)
        cell.font = Font(size=10, bold=(col == 5))
        cell.alignment = Alignment(horizontal="center" if col > 1 else "left", vertical="center")
    ws.row_dimensions[r].height = 20

# Anchos
ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 18
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 18

wb.save("inventory_report.xlsx")
print("\n💾 Report saved as inventory_report.xlsx")